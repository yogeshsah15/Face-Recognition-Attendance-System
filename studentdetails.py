# improved_studentdetails.py
from tkinter import *
from tkinter import ttk, messagebox, filedialog
from PIL import Image, ImageTk
import os
import cv2
import datetime
import numpy as np
from openpyxl import Workbook, load_workbook
import csv  # <-- added for CSV saving

class StudentDetails:
    def __init__(self, root):
        self.root = root
        self.root.title("\U0001F4DA Student Registration Panel")
        self.root.geometry("1100x600")
        self.root.config(bg="#f0f4f7")

        self.var_id = StringVar()
        self.var_name = StringVar()
        self.var_dept = StringVar()
        self.var_sem = StringVar()
        self.var_roll = StringVar()
        self.var_gender = StringVar()
        self.var_email = StringVar()
        self.var_phone = StringVar()
        self.var_photo = StringVar()

        title = Label(self.root, text="\U0001F393 Student Details Management", font=("Helvetica", 20, "bold"), bg="#2c3e50", fg="white", pady=10)
        title.pack(fill=X)

        entry_frame = Frame(self.root, bd=2, relief=RIDGE, bg="white")
        entry_frame.place(x=10, y=70, width=500, height=500)

        lbl_title = Label(entry_frame, text="Enter Student Information", font=("Arial", 14, "bold"), bg="white", fg="#2c3e50")
        lbl_title.grid(row=0, columnspan=2, pady=10)

        labels = ["Student ID", "Name", "Department", "Semester", "Roll No", "Gender", "Email", "Phone"]
        for i, label in enumerate(labels):
            Label(entry_frame, text=f"{label}:", font=("Arial", 12, "bold"), bg="white", anchor="w").grid(row=i+1, column=0, sticky=W, padx=10, pady=5)

        entries = [
            Entry(entry_frame, textvariable=self.var_id, font=("Arial", 12)),
            Entry(entry_frame, textvariable=self.var_name, font=("Arial", 12)),
            Entry(entry_frame, textvariable=self.var_dept, font=("Arial", 12)),
            Entry(entry_frame, textvariable=self.var_sem, font=("Arial", 12)),
            Entry(entry_frame, textvariable=self.var_roll, font=("Arial", 12)),
            ttk.Combobox(entry_frame, textvariable=self.var_gender, font=("Arial", 12), values=["Male", "Female", "Other"], state="readonly"),
            Entry(entry_frame, textvariable=self.var_email, font=("Arial", 12)),
            Entry(entry_frame, textvariable=self.var_phone, font=("Arial", 12)),
        ]
        for i, entry in enumerate(entries):
            entry.grid(row=i+1, column=1, padx=10, pady=5, sticky=W)

        Button(entry_frame, text="\U0001F4C1 Upload Photo", command=self.upload_photo, font=("Arial", 11), bg="#2980b9", fg="white").grid(row=9, column=0, padx=10, pady=10, sticky=W)
        Button(entry_frame, text="\U0001F4F8 Take Photo", command=self.take_photo, font=("Arial", 11), bg="#27ae60", fg="white").grid(row=9, column=1, padx=10, pady=10, sticky=W)

        btn_frame = Frame(entry_frame, bg="white")
        btn_frame.grid(row=10, columnspan=2, pady=20)

        Button(btn_frame, text="\U0001F4BE Save", width=10, font=("Arial", 11), bg="#2ecc71", fg="white", command=self.save_data).grid(row=0, column=0, padx=5)
        Button(btn_frame, text="\U0001F501 Update", width=10, font=("Arial", 11), bg="#f39c12", fg="white").grid(row=0, column=1, padx=5)
        Button(btn_frame, text="\U0001F5D1 Delete", width=10, font=("Arial", 11), bg="#e74c3c", fg="white").grid(row=0, column=2, padx=5)
        Button(btn_frame, text="\U0001F504 Reset", width=10, font=("Arial", 11), bg="#7f8c8d", fg="white").grid(row=0, column=3, padx=5)

        table_frame = Frame(self.root, bd=2, relief=RIDGE, bg="white")
        table_frame.place(x=520, y=70, width=570, height=500)

        Label(table_frame, text="\U0001F50D Search by ID:", font=("Arial", 12, "bold"), bg="white").pack(side=TOP, anchor="w", padx=10, pady=5)
        self.search_var = StringVar()
        Entry(table_frame, textvariable=self.search_var, font=("Arial", 12)).pack(side=TOP, anchor="w", padx=10)
        Button(table_frame, text="Search", font=("Arial", 11), bg="#3498db", fg="white").pack(side=TOP, anchor="w", padx=10, pady=5)

        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Arial", 12, "bold"))
        style.configure("Treeview", font=("Arial", 11), rowheight=28)
        style.map("Treeview", background=[('selected', '#3498db')], foreground=[('selected', 'white')])

        self.student_table = ttk.Treeview(table_frame, columns=("id", "name", "dept", "sem", "roll", "gender", "email", "phone"), show='headings')
        headings = ["Student ID", "Name", "Department", "Semester", "Roll No", "Gender", "Email", "Phone"]
        for col, head in zip(self.student_table["columns"], headings):
            self.student_table.heading(col, text=head)
            self.student_table.column(col, width=80 if col == "id" else 100)

        self.student_table.pack(fill=BOTH, expand=1, padx=10, pady=10)

    def upload_photo(self):
        path = filedialog.askopenfilename(title="Select Photo", filetypes=[("Image Files", "*.jpg *.png *.jpeg")])
        if path:
            self.var_photo.set(path)
            messagebox.showinfo("Success", "Photo uploaded successfully!")

    def take_photo(self):
        if not self.var_id.get() or not self.var_name.get():
            messagebox.showwarning("Missing Info", "Enter Student ID and Name first.")
            return

        cam = cv2.VideoCapture(0)
        face_detector = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
        count = 0
        os.makedirs("photos", exist_ok=True)

        while True:
            ret, img = cam.read()
            if not ret:
                break

            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = face_detector.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5)

            for (x, y, w, h) in faces:
                count += 1
                face_img = gray[y:y+h, x:x+w]
                file_name = f"photos/user.{self.var_id.get()}.{count}.jpg"
                cv2.imwrite(file_name, face_img)

                cv2.rectangle(img, (x, y), (x+w, y+h), (0, 255, 0), 2)
                cv2.putText(img, f"Capturing {count}/20", (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)

            cv2.imshow("\U0001F4F8 Capturing Face", img)

            if cv2.waitKey(1) == 27 or count >= 20:
                break

        cam.release()
        cv2.destroyAllWindows()

        if count >= 20:
            messagebox.showinfo("Success", "\u2705 20 face images captured successfully!")
            train_classifier("photos")
        else:
            messagebox.showwarning("Incomplete", f"Only {count} images captured. Try again.")

    def save_data(self):
        if not self.var_id.get() or not self.var_name.get():
            messagebox.showwarning("Incomplete Data", "Please fill all required fields")
            return

        data = (
            self.var_id.get(), self.var_name.get(), self.var_dept.get(),
            self.var_sem.get(), self.var_roll.get(), self.var_gender.get(),
            self.var_email.get(), self.var_phone.get()
        )

        # Insert into table
        self.student_table.insert("", END, values=data)

        # Save to CSV
        csv_file = "studentdetails.csv"
        file_exists = os.path.isfile(csv_file)
        with open(csv_file, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if not file_exists:
                writer.writerow(["Student ID", "Name", "Department", "Semester", "Roll No", "Gender", "Email", "Phone"])
            writer.writerow(data)

        # Save to attendance.xlsx
        file = "attendance.xlsx"
        now = datetime.datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")

        if not os.path.exists(file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance"
            ws.append(["Student ID", "Name", "Date", "Time"])
        else:
            wb = load_workbook(file)
            ws = wb.active

        ws.append([self.var_id.get(), self.var_name.get(), date_str, time_str])
        wb.save(file)
        messagebox.showinfo("Success", "Student and attendance data saved successfully!")

def train_classifier(data_dir="photos"):
    recognizer = cv2.face.LBPHFaceRecognizer_create()
    detector = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
    faces = []
    ids = []
    for img_name in os.listdir(data_dir):
        if img_name.lower().endswith(('.jpg', '.jpeg', '.png')):
            img_path = os.path.join(data_dir, img_name)
            gray_img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)
            if gray_img is None:
                continue
            try:
                id = int(img_name.split('.')[1])
            except:
                continue
            faces.append(gray_img)
            ids.append(id)
    if len(set(ids)) < 2:
        messagebox.showwarning("Training Skipped", "At least 2 different student faces required to train model.")
        return
    recognizer.train(faces, np.array(ids))
    recognizer.save("classifier.xml")
    messagebox.showinfo("Training Complete", "\U0001F389 Model trained and saved as classifier.xml")

if __name__ == "__main__":
    root = Tk()
    obj = StudentDetails(root)
    root.mainloop()
