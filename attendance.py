from tkinter import *
from tkinter import ttk, messagebox
import os
from openpyxl import load_workbook  # For reading Excel files

class Attendance:
    def __init__(self, root):
        self.root = root
        self.root.title("Attendance Records")
        self.root.geometry("700x400")
        self.root.config(bg="white")

        title = Label(self.root, text="📋 Attendance Report (.xlsx)", font=("Helvetica", 16, "bold"), bg="white", fg="#333")
        title.pack(pady=10)

        # Define treeview
        self.table = ttk.Treeview(self.root, columns=("id", "name", "date", "time"), show='headings')
        self.table.heading("id", text="Student ID")
        self.table.heading("name", text="Name")
        self.table.heading("date", text="Date")
        self.table.heading("time", text="Time")

        self.table.column("id", width=80)
        self.table.column("name", width=200)
        self.table.column("date", width=100)
        self.table.column("time", width=100)

        self.table.pack(fill=BOTH, expand=1, padx=20, pady=10)

        self.load_data()

    def load_data(self):
        filename = "attendance.xlsx"
        if not os.path.exists(filename):
            messagebox.showwarning("No File", f"{filename} not found.")
            return

        try:
            wb = load_workbook(filename)
            sheet = wb.active
            self.table.delete(*self.table.get_children())  # Clear previous data

            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
                if row and len(row) >= 4:
                    self.table.insert("", END, values=(row[0], row[1], row[2], row[3]))
        except Exception as e:
            messagebox.showerror("Read Error", f"Failed to read Excel file.\n{str(e)}")

if __name__ == "__main__":
    root = Tk()
    obj = Attendance(root)
    root.mainloop()
