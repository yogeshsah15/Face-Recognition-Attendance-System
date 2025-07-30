import cv2
import numpy as np
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

def mark_attendance(student_id, name):
    filename = "attendance.xlsx"
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    if not os.path.exists(filename):
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Student ID", "Name", "Date", "Time"])
        wb.save(filename)

    wb = load_workbook(filename)
    sheet = wb.active
    sheet.append([student_id, name, date_str, time_str])
    wb.save(filename)

# Load the trained face classifier
recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read("trainer.yml")

face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

# Load your student info (map ID to Name)
# In production, this should be loaded from a DB or studentdetails.csv
id_name_map = {}
with open("studentdetails.csv", "r") as f:
    lines = f.readlines()[1:]  # skip header
    for line in lines:
        parts = line.strip().split(",")
        if len(parts) >= 2:
            sid, name = parts[0], parts[1]
            id_name_map[int(sid)] = name

cap = cv2.VideoCapture(0)

while True:
    ret, frame = cap.read()
    if not ret:
        break

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, 1.3, 5)

    for (x, y, w, h) in faces:
        face_roi = gray[y:y+h, x:x+w]
        id_, conf = recognizer.predict(face_roi)

        if conf < 60:  # Confidence threshold
            name = id_name_map.get(id_, "Unknown")
            mark_attendance(id_, name)
            label = f"{name} ({round(conf, 2)}%)"
            color = (0, 255, 0)
        else:
            label = "Unknown"
            color = (0, 0, 255)

        cv2.rectangle(frame, (x, y), (x + w, y + h), color, 2)
        cv2.putText(frame, label, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, color, 2)

    cv2.imshow("Face Recognition Attendance", frame)

    if cv2.waitKey(1) == 27:  # ESC to exit
        break

cap.release()
cv2.destroyAllWindows()
